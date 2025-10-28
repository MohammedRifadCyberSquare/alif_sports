import base64
from io import BytesIO
import io
import base64
import eel
import matplotlib.pyplot as plt
from xhtml2pdf import pisa
from jinja2 import Environment, FileSystemLoader
from jinja2 import Template
from openpyxl import load_workbook
import ctypes
import os
from xlsx2html import xlsx2html
import pdfkit
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy import func, desc
from datetime import datetime
import win32com.client
from xlsx2html import xlsx2html
from constants import ITEMS

from web.models.db_models import Base, engine, SessionLocal, Student, ParticipantItem, Result, House, ResultGrp

eel.init('web')  # Frontend folder

# Ensure the database is created
Base.metadata.create_all(engine)

# Create session factory
session = SessionLocal()

user32 = ctypes.windll.user32
screen_width = user32.GetSystemMetrics(0)
screen_height = user32.GetSystemMetrics(1)
PDF_FOLDER = os.path.join(os.getcwd(), 'web/pdfs')
width = int(screen_width * 1)
height = int(screen_height * 1)

x = int((screen_width - width) / 2)
y = int((screen_height - height) / 2)

# Set environment variable for position (works on Windows)
os.environ['EEL_START_POSITION'] = f"{x},{y}"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
eel.start('index.html', size=(width, height), block=False)

import sys
import os

if getattr(sys, 'frozen', False):
    # Running as bundled exe
    base_path = sys._MEIPASS
else:
    base_path = os.path.abspath(".")

def resource_path(relative_path):
    return os.path.join(base_path, relative_path)


html_path = resource_path("templates")   
db_path = resource_path("sports_fest.db")



@eel.expose
def get_categories():
    """Return list of categories"""
    return list(ITEMS.keys())


@eel.expose
def get_items(category):
    """Return list of items for a given category"""
    if category.lower() in ITEMS:
        print(
            f"Items for category {category}: {[x['item'] for x in ITEMS[category.lower()]]}")
        return [x['item'] for x in ITEMS[category.lower()]]
    return []


@eel.expose
def register_student(base64_file):
    session = SessionLocal()

    try:
        # Decode base64 to bytes
        file_bytes = base64.b64decode(base64_file)
        workbook = load_workbook(filename=BytesIO(file_bytes))
        sheet = workbook.active

        chest_no = 0

        rows_added = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            clean_row = [str(cell).strip()
                         if cell is not None else None for cell in row]
            row = list(row[1:])
            print(row)

            admission_no, cls, student_name, dob, division, \
                category, house = row[:7]

            if isinstance(dob, datetime):
                dob_str = dob.strftime("%d/%m/%Y")
            else:
                dob_str = str(dob).strip() if dob else None

            last_chest = session.query(func.max(Student.chest_no)).filter(
                Student.house == house.lower(),
            ).scalar()

            if not last_chest:
                if house.lower() == 'alpha':
                    chest_no = 1000
                elif house.lower() == 'beta':
                    chest_no = 2000
                elif house.lower() == 'gamma':
                    chest_no = 3000
                elif house.lower() == 'delta':
                    chest_no = 4000
            else:
                chest_no = last_chest + 1

            student = Student(
                admission_no=admission_no,
                student_name=student_name.lower(),
                student_class=cls,
                division=division.lower(),
                dob=dob_str,
                category=category.lower(),
                house=house.lower(),
                chest_no=chest_no,
            )

            session.add(student)
            session.flush()
            rows_added += 1
            session.commit()
        return {"status": "success", "message": f"{rows_added} participants added successfully!"}
    except Exception as e:
        print("Error:", e)
        session.rollback()
        return {"status": "error", "message": str(e)}


@eel.expose
def fetch_students_for_event_registeration(category, team):
    session = SessionLocal()
    try:
        students = session.query(Student).filter(
            Student.category == category.lower(),
            Student.house == team
        ).order_by(Student.chest_no).all()

        students_data = [{
            'admission_no': s.admission_no,
            'student_name': s.student_name,
            'chest_no': s.chest_no,
        } for s in students]

        return {"statusCode": 200, "students": students_data}
    except Exception as e:
        print("Error:", e)
        return {"status": "error", "message": str(e)}


def generate_pdf_from_html(template_path, output_path, context):
    env = Environment(loader=FileSystemLoader(os.path.join(BASE_DIR, 'web')))
    template = env.get_template(template_path)
    html_content = template.render(context)

    with open(output_path, 'wb') as f:
        pisa.CreatePDF(html_content, dest=f)

    print(f"!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: {output_path}")
    # Return output path so frontend can open
    return output_path


@eel.expose
def get_team_list(category, team):
    session = SessionLocal()
    print(f"Generating team list for Category: {category}, Team: {team}")
    try:
        students = session.query(Student).filter(
            Student.category == category.lower(),
            Student.house == team.lower()
        ).order_by(Student.chest_no).all()

        print(f"Found {len(students)} students")

        pdf_filename = f"{team}_{category}.pdf"
        pdf_path = os.path.join(PDF_FOLDER, pdf_filename)
        os.makedirs(PDF_FOLDER, exist_ok=True)
        logo_path = os.path.abspath(os.path.join(
            "web", "template", "images", "logo1.png"))
        template_path = 'report_templates/team_list.html'

        students_data = [{
            'admission_no': s.admission_no,
            'student_name': s.student_name,
            'chest_no': s.chest_no,
            'dob': s.dob,
        } for s in students]

        generate_pdf_from_html(
            template_path,
            pdf_path,
            {
                'category': category.upper(),
                'team': team.upper(),
                'students': students,
                'logo_path': logo_path
            }
        )

        return {"statusCode": 200, "students": students_data, "pdfPath": f"pdfs/{pdf_filename}"}
    except Exception as e:
        print("Error:", e)
        return {"status": "error", "message": str(e)}


@eel.expose
def get_team_cat_list(team, category, item):
    try:
        students = (
            session.query(Student)
            .join(ParticipantItem, ParticipantItem.participant_id == Student.admission_no)
            .filter(
                ParticipantItem.category == category.lower(),
                ParticipantItem.item == item.lower(),
                Student.house == team.lower()
            )
            .order_by(Student.chest_no)
            .all()
        )

        print(f"Found {len(students)} students")

        pdf_filename = f"{team}_{category}.pdf"
        pdf_path = os.path.join(PDF_FOLDER, pdf_filename)
        os.makedirs(PDF_FOLDER, exist_ok=True)

        logo_path = os.path.abspath(os.path.join("web", "template", "images", "logo1.png"))
        template_path = "report_templates/team_cat_item_list.html"

        students_data = [
            {
                "admission_no": s.admission_no,
                "student_name": s.student_name,
                "chest_no": s.chest_no,
                "dob": s.dob,
            }
            for s in students
        ]

        generate_pdf_from_html(
            template_path,
            pdf_path,
            {
                "category": category.upper(),
                "team": team.upper(),
                "students": students_data,
                "logo_path": logo_path,
            },
        )

        return {
            "statusCode": 200,
            "students": students_data,
            "pdfPath": f"pdfs/{pdf_filename}",
        }

    except Exception as e:
        print("Error:", e)
        return {"status": "error", "message": str(e)}



@eel.expose
def check_student_item_count(admission_no, type):
    record_exist_count = session.query(ParticipantItem).filter_by(
        participant_id=admission_no,
        type=type  # Use participant.id, not admission_no

    ).count()
    if record_exist_count >= 3:
        return {"statusCode": 409, "message": f"Admission No {admission_no} has already filled the capacity"}
    else:
        return {"statusCode": 200, "message": "ok"}


@eel.expose
def register_students(category, item, type, students):
    print(category, item, type, students)
    session = SessionLocal()
    item_count = 0

    for admission_no in students:

        participant = session.query(Student).filter_by(
            admission_no=admission_no).first()

        record = ParticipantItem(
            participant_id=participant.admission_no,
            category=category,
            item=item,
            type=type
        )

        session.add(record)
        session.flush()
        session.commit()

        return {"statusCode": 201, "message": "Registered Succesfully"}

@eel.expose
def highest_score_by_category():
    session = SessionLocal()
    try:
        categories = ["junior", "sub junior", "senior"]
        top_scorers = []

        for cat in categories:
            # Fetch top scorer per category
            student = (
                session.query(Student)
                .filter(Student.category == cat)
                .order_by(Student.points.desc())
                .first()
            )

            if student:
                top_scorers.append({
                    "category": cat.upper(),
                    "admission_no":student.admission_no,
                    "chest_no":student.chest_no,
                    "student_name": student.student_name.upper(),
                    "house": student.house.upper(),
                    "total_points": student.points
                })

        if not top_scorers:
            return {"statusCode": 404, "message": "No students found"}

        print(top_scorers,'*************')
        # ---------- PDF generation ----------
        pdf_filename = "highest_scorers.pdf"
        pdf_path = os.path.join(PDF_FOLDER, pdf_filename)
        os.makedirs(PDF_FOLDER, exist_ok=True)

        logo_path = os.path.abspath(os.path.join(
            "web", "template", "images", "logo1.png"))
        template_path = "report_templates/highest.html"

        # Context for template
        context = {
            "title": "Highest Scorers by Category",
            "logo_path": logo_path,
            "scorers": top_scorers,
            "date": datetime.now().strftime("%d %B %Y")
        }

        generate_pdf_from_html(template_path, pdf_path, context)

        return {
            "statusCode": 200,
            "topScorers": top_scorers,
            "pdfPath": f"pdfs/{pdf_filename}"
        }

    except Exception as e:
        print("Error fetching top scorers:", e)
        return {"statusCode": 500, "message": str(e)}

    finally:
        session.close()

@eel.expose
def add_result(category, item, type, position1="", position2="", position3=""):
    print('********************************', category,
          item, type, position1, position2, position3)

    session = SessionLocal()
    try:
        # Normalize inputs
        category = category.strip().lower()
        item = item.strip().lower()
        type = type.strip().lower()

        # Delete existing results for this event
        session.query(Result).filter_by(category=category, item=item).delete()
        session.query(ResultGrp).filter_by(
            category=category, item=item).delete()

        results = []

        # ---------- GROUP RESULTS ----------
        if type == "group":
            print("Adding group results...")
            if position1:
                results.append(ResultGrp(
                    category=category,
                    item=item,
                    house_name=position1,
                    position="first"
                ))
                # house = session.query(House).filter(
                #     House.house_name ==position1.lower()
                # ).first()
                # house.total_points+=10

            if position2:
                results.append(ResultGrp(
                    category=category,
                    item=item,
                    house_name=position2,
                    position="second"
                ))

                # house = session.query(House).filter(
                #     House.house_name ==position2.lower()
                # ).first()
                # house.total_points+=5

        # ---------- INDIVIDUAL RESULTS ----------
        else:
            print("Adding individual results...")
            if position1:
                results.append(Result(
                    category=category,
                    item=item,
                    type=type,
                    participant_id=position1,
                    position="first"
                ))
            if position2:
                results.append(Result(
                    category=category,
                    item=item,
                    type=type,
                    participant_id=position2,
                    position="second"
                ))
            if position3:
                results.append(Result(
                    category=category,
                    item=item,
                    type=type,
                    participant_id=position3,
                    position="third"
                ))

        if results:
            session.add_all(results)
            session.commit()
            print(f"✅ Saved {len(results)} results successfully.")
            return {"statusCode": 201, "message": "Result Added Successfully"}
        else:
            print("⚠️ No results to save.")
            return {"statusCode": 400, "message": "No positions provided."}

    except Exception as e:
        session.rollback()
        print("❌ Error in add_result:", e)
        return {"statusCode": 500, "message": str(e)}

    finally:
        session.close()

@eel.expose
def get_result_by_category_and_item(type,category, item):
    print(type,category, item)
    try:
        if type == "individual":

            records = (
                session.query(Result, Student)
                .join(Student, Result.participant_id == Student.admission_no)
                .filter(
                    Result.item == item.lower(),
                    Result.category == category.lower()
                )
                .all()
            )

            students_data = [
                {
                    "student_name": s.student_name,
                    "chest_no": s.chest_no,
                    "admission_no":s.admission_no,
                    "house": s.house,
                    "position": r.position,
                    "category": r.category,
                    "item": r.item
                }
                for r, s in records
            ]

            pdf_filename = f"{item}_{category}.pdf"
            pdf_path = os.path.join(PDF_FOLDER, pdf_filename)
            os.makedirs(PDF_FOLDER, exist_ok=True)
            logo_path = os.path.abspath(os.path.join(
                "web", "template", "images", "logo1.png"))
            template_path = 'report_templates/result_cat_item_ind.html'


            generate_pdf_from_html(
                template_path,
                pdf_path,
                {
                    "category": category.upper(),
                    "students": students_data,
                    "logo_path": logo_path,
                    "item":item.upper()
                },
            )
            print(pdf_path)
            return {
                "statusCode": 200,
                "students": students_data,
                "pdfPath": f"pdfs/{pdf_filename}",
            }

            

        else:
            print('heeeeeeeeeeeeeeeeeeeeeeere')
            print(item,category)
            records = (
                session.query(ResultGrp)
                .filter(
                    ResultGrp.item == item.lower(),
                    ResultGrp.category == category.lower()
                )
                .all()
            )

            result_data = [
                {
                    "house": r.house_name,
                    "position": r.position,
                    "category": r.category,
                     "item":item.upper()
                }
                for r in records
            ]

            print(result_data)


            pdf_filename = f"{item}_{category}.pdf"
            pdf_path = os.path.join(PDF_FOLDER, pdf_filename)
            os.makedirs(PDF_FOLDER, exist_ok=True)
            logo_path = os.path.abspath(os.path.join(
                "web", "template", "images", "logo1.png"))
            template_path = 'report_templates/result_cat_item_grp.html'


            generate_pdf_from_html(
                template_path,
                pdf_path,
                {
                    "category": category.upper(),
                    "students": result_data,
                    "logo_path": logo_path,
                    "item":item.upper()
                },
            )

            return {
                "statusCode": 200,
                "students": result_data,
                "pdfPath": f"pdfs/{pdf_filename}",
            }


    except Exception as e:
        print(e,'//////////////////')





@eel.expose
def get_pending_grp_result():
    records = session.query(ResultGrp).filter(
        ResultGrp.is_finalised == 0
    ).all()
    grouped_data = {}
    for r in records:
        key = (r.category, r.item)
        if key not in grouped_data:
            grouped_data[key] = {
                "category": r.category.upper(),
                "item": r.item.upper(),
                "first": "",
                "second": "",
                "third": ""
            }

        # Assign based on position
        if "first" in str(r.position):
            grouped_data[key]["first"] = r.house_name.upper()
        elif "second" in str(r.position):
            grouped_data[key]["second"] = r.house_name.upper()

    return {"statusCode": 200, "data": list(grouped_data.values())}


@eel.expose
def finalise_grp_result(selected_list):
    print(selected_list)
    for item in selected_list:
        cat = item["category"].lower()
        selected_item = item["item"].lower()
        first = item["first"].lower()
        second = item["second"].lower()
        print(cat, item, first, second)

        record_first = session.query(ResultGrp).filter(
            ResultGrp.category == cat,
            ResultGrp.item == selected_item,
            ResultGrp.house_name == first
        ).first()

        record_first.is_finalised = 1

        house_record_first = session.query(House).filter(
            House.house_name == first
        ).first()

        house_record_first.total_points += 10

        record_second = session.query(ResultGrp).filter(
            ResultGrp.category == cat,
            ResultGrp.item == selected_item,
            ResultGrp.house_name == second
        ).first()

        record_second.is_finalised = 1

        house_record_second = session.query(House).filter(
            House.house_name == second
        ).first()

        house_record_second.total_points += 5

        session.commit()

    return {"statusCode": 200, }


@eel.expose
def get_house_points():

    houses = session.query(House).all()
    result = [{"house": h.house_name, "points": h.total_points}
              for h in houses]

    if not result:
        return {"statusCode": 404, "message": "No houses found"}

    # Sort by points (descending)
    sorted_result = sorted(result, key=lambda x: x["points"], reverse=True)
    winner = sorted_result[0]

    # Define colors for each house
    house_colors = {
        "alpha": "#ff4d4d",  # red
        "beta": "#3399ff",   # blue
        "gamma": "#33cc33",  # green
        "delta": "#ffcc00"   # yellow
    }

    # Prepare chart data
    names = [r["house"] for r in sorted_result]
    points = [r["points"] for r in sorted_result]
    colors = [house_colors.get(name.lower(), "#808080")
              for name in names]  # default gray

    # Create chart
    plt.figure(figsize=(8, 5))
    bars = plt.bar(names, points, color=colors)
    plt.title("🏆 House Points", fontsize=16)
    plt.xlabel("House", fontsize=12)
    plt.ylabel("Points", fontsize=12)

    # Show values on bars
    for bar in bars:
        height = bar.get_height()
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            height + 0.2,
            f'{int(height)}',
            ha='center',
            va='bottom',
            fontsize=10,
            color='black',
            fontweight='bold'
        )

    # Winner label
    plt.figtext(
        0.5, -0.05,
        f"🏅 Winner: {winner['house'].capitalize()} ({winner['points']} points)",
        ha='center',
        fontsize=12,
        fontweight='bold',
        color='green'
    )

    plt.tight_layout()

    # Convert chart to base64
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    image_base64 = base64.b64encode(buf.read()).decode('utf-8')
    plt.close()
    return {"statusCode": 200, "imgData": image_base64, "winner": winner}
    # return image_base64


@eel.expose
def get_pending_results():

    records = (
        session.query(Result, Student.student_name)
        .join(Student, Student.admission_no == Result.participant_id)
        .filter(Result.is_finalised == 0)
        .all()
    )
    grouped = {}
    for r, student_name in records:
        key = (r.category, r.item, r.type)

        if key not in grouped:
            grouped[key] = {"category": r.category.upper(), "item": r.item.upper(), "type": r.type.upper(),
                            "first": None, "second": None, "third": None}

        student_display = f"{student_name.upper()} ({r.participant_id})"
        if r.position.lower() == "first":
            grouped[key]["first"] = student_display
        elif r.position.lower() == "second":
            grouped[key]["second"] = student_display
        elif r.position.lower() == "third":
            grouped[key]["third"] = student_display

    result = list(grouped.values())

    print(result)
    return {"statusCode": 200, "data": result}


@eel.expose
def add_participant(team, cat, adm_no, s_name, grade, div, dob):
    if isinstance(dob, datetime):
        dob_str = dob.strftime("%d/%m/%Y")
    else:
        dob_str = str(dob).strip() if dob else None

    record = session.query(Student).filter(
        Student.admission_no == adm_no).first()

    if record:
        return {"statusCode": 409, "message": "Student Already Added"}

    last_chest = (
        session.query(func.max(Student.chest_no))
        .filter(Student.house == team)
        .scalar()
    )

    # Assign next chest number
    new_chest = (last_chest or 0) + 1
    record = Student(
        admission_no=adm_no,
        student_name=s_name,
        student_class=grade,
        division=div,
        dob=dob,
        category=cat,
        house=team,
        chest_no=new_chest
    )
    session.add(record)
    session.flush()
    session.commit()
    return {"statusCode": 201, "message": "Student Registered Succesfully"}


@eel.expose
def finalise_result(selected_records):
    print('*********************', selected_records)

    for record in selected_records:
        print('-----------------------', record, '---------------------------')
        rows = session.query(Result).filter(
            Result.category == record["category"].lower(),
            Result.item == record["item"].lower()
        ).all()
        print(rows)
        for row in rows:
            print('^^^^^^^^^^^^^^^^^^^^^^^^^^^^', row,
                  '^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^')

            student = session.query(Student).filter(
                Student.admission_no == row.participant_id).first()

            house = student.house

            if not student:
                continue
            if row.type == "group":
                group_record = session.query(House).filter(
                    House.house_name == house
                ).first()
                if row.position.lower() == "first":
                    group_record.total_points += 10

                if row.position.lower() == "second":
                    group_record.total_points += 5
            else:
                house_record = session.query(House).filter(
                    House.house_name == house
                ).first()
                if row.position.lower() == "first":
                    student.points += 5
                    house_record.total_points += 5

                elif row.position.lower() == "second":
                    student.points += 3
                    house_record.total_points += 3

                elif row.position.lower() == "third":
                    student.points += 1
                    house_record.total_points += 1

            # Mark result as finalised
            row.is_finalised = 1
    session.commit()
    return {"statusCode": 200, "message": "Details Updated Succesfully"}


while True:
    eel.sleep(1.0)
